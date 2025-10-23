from flask import Flask, render_template, request, send_file, jsonify
import pdfplumber
import pandas as pd
import re
import os
from werkzeug.utils import secure_filename
from datetime import datetime
import io

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'

# Create necessary folders
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

def clean_text(text):
    """Remove illegal characters for Excel"""
    if not isinstance(text, str):
        return text
    return ''.join(char if ord(char) >= 32 or char in '\n\r\t' else ' ' for char in text)

def clean_currency(value):
    """Remove rupee symbol and return clean number string"""
    if not value or value == '':
        return ''
    return str(value).replace('₹', '').strip()

def extract_uber_transactions(pdf_path):
    """Extract transaction data from Uber statement PDF with complete descriptions"""
    
    all_transactions = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            text = page.extract_text()
            if not text:
                continue
            
            lines = text.split('\n')
            i = 0
            
            while i < len(lines):
                line = lines[i].strip()
                
                # Skip headers and footers
                if ('Praveen K -' in line or 'Weekly Statement' in line or 
                    'Processed Event' in line or line == 'Transactions' or 
                    'Jan 15, 2024' in line or not line):
                    i += 1
                    continue
                
                # Look for transaction starting with date
                date_match = re.match(r'^(Mon|Tue|Wed|Thu|Fri|Sat|Sun),\s+(\w+)\s+(\d+)\s+(.+)', line)
                
                if date_match:
                    date_str = f"{date_match.group(1)}, {date_match.group(2)} {date_match.group(3)}"
                    rest_of_first_line = date_match.group(4)
                    
                    # Initialize variables
                    event_description_parts = []
                    earnings = ""
                    payouts = ""
                    balance = ""
                    time_str = ""
                    
                    # Parse first line to extract event description start and financial values
                    parts = rest_of_first_line.split()
                    
                    for part in parts:
                        if re.match(r'^-?₹[\d,]+\.\d+$', part):
                            if not earnings:
                                earnings = part
                            elif not payouts:
                                payouts = part
                        else:
                            event_description_parts.append(part)
                    
                    # Move to next line
                    i += 1
                    
                    # Continue collecting event description until we hit the "processed date" line
                    while i < len(lines):
                        current_line = lines[i].strip()
                        
                        # Check if we hit next transaction
                        if re.match(r'^(Mon|Tue|Wed|Thu|Fri|Sat|Sun),\s+(\w+)\s+(\d+)', current_line):
                            i -= 1
                            break
                        
                        # Check for footer
                        if 'Praveen K -' in current_line or not current_line:
                            break
                        
                        # Clean null bytes for proper matching
                        current_line_clean = current_line.replace('\x00', ':')
                        
                        # Check if this is the "processed date" line
                        processed_date_match = re.match(r'^[A-Z][a-z]+\s+\d+\s+\d+:\d+\s+(AM|PM)\s*$', current_line_clean)
                        if processed_date_match:
                            break
                        
                        # Check for time at start of line
                        time_match = re.match(r'^(\d+):(\d+)\s+(AM|PM)', current_line_clean)
                        if time_match and not time_str:
                            time_str = time_match.group()
                            
                            # Extract balance from this line
                            rupee_amounts = re.findall(r'₹[\d,]+\.\d+', current_line_clean)
                            if rupee_amounts:
                                balance = rupee_amounts[-1]
                            
                            # Get rest of line after time for event description
                            rest = current_line_clean[time_match.end():].strip()
                            rest = re.sub(r'Jan\s+\d+\s+\d+:\d+\s+(AM|PM)', '', rest)
                            rest = re.sub(r'₹[\d,]+\.\d+', '', rest)
                            rest = rest.strip()
                            
                            if rest:
                                event_description_parts.append(rest)
                        else:
                            # This is event description continuation
                            if current_line and not re.match(r'^\d+[\d,]*\.\d+$', current_line):
                                event_description_parts.append(current_line)
                        
                        i += 1
                    
                    # Combine event description
                    event_description = ' '.join(event_description_parts).strip()
                    
                    # Build full date-time string
                    full_datetime = date_str
                    if time_str:
                        full_datetime += f" {time_str}"
                    
                    # Only add transaction if we have meaningful data
                    if event_description or earnings or balance:
                        transaction = {
                            'Date': full_datetime,
                            'Event Description': event_description,
                            'Your Earnings': earnings,
                            'Payouts': payouts,
                            'Balance': balance
                        }
                        all_transactions.append(transaction)
                    
                    continue
                
                i += 1
    
    return all_transactions

def create_excel(transactions):
    """Create Excel file from transactions and return as bytes"""
    
    # Clean all text fields and remove rupee symbols
    for trans in transactions:
        trans['Date'] = clean_text(trans['Date'])
        trans['Event Description'] = clean_text(trans['Event Description'])
        trans['Your Earnings'] = clean_currency(trans['Your Earnings'])
        trans['Payouts'] = clean_currency(trans['Payouts'])
        trans['Balance'] = clean_currency(trans['Balance'])
    
    df = pd.DataFrame(transactions)
    
    # Create Excel file in memory
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Transactions')
        
        worksheet = writer.sheets['Transactions']
        
        # Adjust column widths
        worksheet.column_dimensions['A'].width = 28
        worksheet.column_dimensions['B'].width = 120
        worksheet.column_dimensions['C'].width = 18
        worksheet.column_dimensions['D'].width = 18
        worksheet.column_dimensions['E'].width = 18
        
        # Format header
        from openpyxl.styles import Font, Alignment, PatternFill
        
        for cell in worksheet[1]:
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Format data rows
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            row[0].alignment = Alignment(horizontal='left', vertical='top')
            row[1].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            for cell in row[2:5]:
                cell.alignment = Alignment(horizontal='right', vertical='top')
    
    output.seek(0)
    return output

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({'error': 'Only PDF files are allowed'}), 400
        
        # Save uploaded file
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(filepath)
        
        # Extract transactions
        transactions = extract_uber_transactions(filepath)
        
        if not transactions:
            os.remove(filepath)
            return jsonify({'error': 'No transactions found in PDF'}), 400
        
        # Create Excel file
        excel_output = create_excel(transactions)
        
        # Clean up uploaded PDF
        os.remove(filepath)
        
        # Generate output filename
        output_filename = f"uber_statement_{timestamp}.xlsx"
        
        return send_file(
            excel_output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=output_filename
        )
    
    except Exception as e:
        return jsonify({'error': f'Error processing file: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
